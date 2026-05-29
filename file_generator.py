"""
File generator for LICIPD pipeline.

Responsibilities:
  1. Append new quarter row to Master_Data/Master_LICIPD_DATA.csv
  2. Generate LICIPD_DATA_YYYYMMDD.xlsx (all historical data)
  3. Generate LICIPD_META_YYYYMMDD.xlsx (measure metadata)
  4. Create LICIPD_YYYYMMDD.zip containing both xlsx files
  5. Copy outputs to output/latest/ and output/<datestamp>/

Public API:
    result = generate(quarter, data)
    # quarter : 'YYYY-QN' string
    # data    : {column_code: value} dict (56 codes)
"""

import os
import csv
import logging
import zipfile
from datetime import datetime

import config

logger = logging.getLogger(__name__)

# META columns in order they appear in the output file
_META_COLS = [
    'CODE', 'CODE_MNEMONIC', 'DESCRIPTION',
    'FREQUENCY', 'MULTIPLIER', 'AGGREGATION_TYPE',
    'UNIT_TYPE', 'DATA_TYPE', 'DATA_UNIT',
    'SEASONALLY_ADJUSTED', 'ANNUALIZED', 'STATE',
    'PROVIDER_MEASURE_URL', 'PROVIDER', 'SOURCE',
    'SOURCE_DESCRIPTION', 'COUNTRY', 'DATASET',
]


# ── Master CSV helpers ────────────────────────────────────────────────────────

def _read_master_csv():
    """
    Read the master CSV and return (header_rows, data_rows).
    header_rows : list of raw row lists (first 2 rows)
    data_rows   : list of raw row lists (data rows only, non-empty)
    """
    header_rows = []
    data_rows   = []

    if not os.path.exists(config.MASTER_CSV):
        logger.warning(f"Master CSV not found: {config.MASTER_CSV}")
        return header_rows, data_rows

    with open(config.MASTER_CSV, newline='', encoding='utf-8-sig') as fh:
        reader = csv.reader(fh)
        for i, row in enumerate(reader):
            if i < 2:
                header_rows.append(row)
            elif any(cell.strip() for cell in row):   # skip blank rows
                data_rows.append(row)

    return header_rows, data_rows


def _existing_quarters(data_rows):
    """Return set of quarter labels already in the master CSV."""
    return {row[0].strip() for row in data_rows if row and row[0].strip()}


def _format_value(val):
    """Format a value for CSV storage: float -> str, 'NA' -> 'NA'."""
    if val == 'NA' or val is None:
        return 'NA'
    if isinstance(val, float):
        # Preserve significant digits without trailing zeros where possible
        return f"{val:g}" if val == int(val) else str(round(val, 6))
    return str(val)


def append_to_master_csv(quarter, data):
    """
    Append a new row to the master CSV.

    Args:
        quarter : 'YYYY-QN' string
        data    : {column_code: value_or_'NA'}

    Returns:
        True if appended, False if quarter already exists.
    """
    header_rows, data_rows = _read_master_csv()
    existing = _existing_quarters(data_rows)

    if quarter in existing:
        logger.info(f"Quarter {quarter} already in master CSV — skipping append")
        return False

    # Build new row in column order
    new_row = [quarter] + [
        _format_value(data.get(code, 'NA'))
        for code in config.COLUMN_CODES
    ]

    # Re-write the file: headers + existing data + new row
    os.makedirs(os.path.dirname(config.MASTER_CSV), exist_ok=True)
    with open(config.MASTER_CSV, 'w', newline='', encoding='utf-8') as fh:
        writer = csv.writer(fh)
        for row in header_rows:
            writer.writerow(row)
        for row in data_rows:
            writer.writerow(row)
        writer.writerow(new_row)

    logger.info(f"Appended {quarter} to master CSV  ({len(data_rows)+1} data rows total)")
    return True


# ── Excel generation ──────────────────────────────────────────────────────────

def _try_numeric(val_str):
    """Try to convert a string to int/float for Excel number cells."""
    if val_str in ('NA', '', None):
        return val_str
    try:
        f = float(val_str)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return val_str


def _write_data_xlsx(path, header_rows, data_rows):
    """Write LICIPD_DATA xlsx: rows = quarters, columns = measures."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl not available — skipping DATA xlsx generation")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LICIPD_DATA'

    hdr_fill   = PatternFill('solid', fgColor='004E8F')
    hdr_font   = Font(bold=True, color='FFFFFF')
    hdr_align  = Alignment(horizontal='center', wrap_text=True)
    label_font = Font(bold=True)

    # Row 1: column codes
    if header_rows:
        for j, val in enumerate(header_rows[0]):
            cell = ws.cell(row=1, column=j+1, value=val)
            if j == 0:
                cell.font = label_font
            else:
                cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, hdr_align

    # Row 2: descriptions
    if len(header_rows) > 1:
        for j, val in enumerate(header_rows[1]):
            cell = ws.cell(row=2, column=j+1, value=val)
            if j > 0:
                cell.alignment = Alignment(wrap_text=True)

    # Data rows
    for i, row in enumerate(data_rows):
        for j, val in enumerate(row):
            cell = ws.cell(row=3+i, column=j+1)
            if j == 0:
                cell.value = val          # quarter label stays as text
                cell.font = label_font
            else:
                cell.value = _try_numeric(val)

    # Auto-width for first column (quarter labels)
    ws.column_dimensions['A'].width = 12
    # Freeze panes below headers, right of quarter column
    ws.freeze_panes = 'B3'

    try:
        wb.save(path)
        logger.info(f"DATA xlsx saved: {path}")
        return True
    except Exception as e:
        logger.error(f"Failed to save DATA xlsx: {e}")
        return False


def _write_meta_xlsx(path):
    """Write LICIPD_META xlsx: 56 rows, one per measure."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl not available — skipping META xlsx generation")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LICIPD_META'

    hdr_fill  = PatternFill('solid', fgColor='004E8F')
    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_align = Alignment(horizontal='center')

    # Header row
    for j, col_name in enumerate(_META_COLS):
        cell = ws.cell(row=1, column=j+1, value=col_name)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, hdr_align

    # One row per measure
    for i, (code, desc) in enumerate(zip(config.COLUMN_CODES, config.COLUMN_DESCRIPTIONS)):
        # CODE_MNEMONIC: strip the "LICIPD." prefix
        mnemonic = code[len('LICIPD.'):] if code.startswith('LICIPD.') else code
        row_data = {
            'CODE':         code,
            'CODE_MNEMONIC': mnemonic,
            'DESCRIPTION':  desc,
        }
        row_data.update(config.META_TEMPLATE)
        for j, col_name in enumerate(_META_COLS):
            val = row_data.get(col_name, '')
            ws.cell(row=2+i, column=j+1, value=val)

    ws.freeze_panes = 'A2'

    try:
        wb.save(path)
        logger.info(f"META xlsx saved: {path}")
        return True
    except Exception as e:
        logger.error(f"Failed to save META xlsx: {e}")
        return False


def _create_zip(zip_path, *file_paths):
    """Create a ZIP archive from the given files. Returns True on success."""
    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fp in file_paths:
                if fp and os.path.exists(fp):
                    zf.write(fp, os.path.basename(fp))
        logger.info(f"ZIP created: {zip_path}")
        return True
    except Exception as e:
        logger.error(f"Failed to create ZIP: {e}")
        return False


def _copy_file(src, dst_dir):
    """Copy src to dst_dir, creating dir if needed. Returns dest path or None."""
    try:
        import shutil
        os.makedirs(dst_dir, exist_ok=True)
        dst = os.path.join(dst_dir, os.path.basename(src))
        shutil.copy2(src, dst)
        return dst
    except Exception as e:
        logger.error(f"Failed to copy {src} -> {dst_dir}: {e}")
        return None


# ── Public entry point ────────────────────────────────────────────────────────

def generate(quarter, data):
    """
    Append new quarter to master CSV and generate output files.

    Args:
        quarter : 'YYYY-QN' string (e.g. '2026-Q1')
        data    : {column_code: value_or_'NA'} for all 56 codes

    Returns dict:
        {
            'appended'   : bool,           # was CSV updated?
            'data_xlsx'  : path or None,
            'meta_xlsx'  : path or None,
            'zip'        : path or None,
            'latest_dir' : path,
        }
    """
    result = {
        'appended':   False,
        'data_xlsx':  None,
        'meta_xlsx':  None,
        'zip':        None,
        'latest_dir': os.path.join(config.OUTPUT_DIR, 'latest'),
    }

    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')

    # ── 1. Append to master CSV ──────────────────────────────────────────────
    appended = append_to_master_csv(quarter, data)
    result['appended'] = appended

    # ── 2. Read master CSV for xlsx generation ───────────────────────────────
    header_rows, data_rows = _read_master_csv()

    # ── 3. Generate DATA xlsx in a temp run folder ───────────────────────────
    run_dir = os.path.join(config.OUTPUT_DIR, date_str)
    os.makedirs(run_dir, exist_ok=True)

    data_filename = f"{config.OUTPUT_DATA_PREFIX}{date_str}.xlsx"
    meta_filename = f"{config.OUTPUT_META_PREFIX}{date_str}.xlsx"
    zip_filename  = f"{config.OUTPUT_ZIP_PREFIX}{date_str}.zip"

    data_path = os.path.join(run_dir, data_filename)
    meta_path = os.path.join(run_dir, meta_filename)
    zip_path  = os.path.join(run_dir, zip_filename)

    if _write_data_xlsx(data_path, header_rows, data_rows):
        result['data_xlsx'] = data_path
    if _write_meta_xlsx(meta_path):
        result['meta_xlsx'] = meta_path

    # ── 4. Create ZIP ────────────────────────────────────────────────────────
    if result['data_xlsx'] or result['meta_xlsx']:
        if _create_zip(zip_path, result['data_xlsx'], result['meta_xlsx']):
            result['zip'] = zip_path

    # ── 5. Copy all outputs to output/latest/ ────────────────────────────────
    latest_dir = result['latest_dir']
    os.makedirs(latest_dir, exist_ok=True)

    for src in [result['data_xlsx'], result['meta_xlsx'], result['zip']]:
        if src and os.path.exists(src):
            _copy_file(src, latest_dir)

    logger.info(
        f"generate() complete: quarter={quarter}  appended={appended}  "
        f"outputs in {run_dir}"
    )
    return result
